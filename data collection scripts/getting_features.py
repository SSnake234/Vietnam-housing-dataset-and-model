import os
import random
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import pandas as pd
from openpyxl import Workbook, load_workbook

# Set up Chromedriver and Selenium
chrome_driver_path = 'G:/chromedriver/chromedriver-win64/chromedriver.exe'

# List of user agents
user_agents = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:85.0) Gecko/20100101 Firefox/85.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:87.0) Gecko/20100101 Firefox/87.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 11_2_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.114 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36"
]

def get_random_user_agent():
    return random.choice(user_agents)

def get_url(url):
    # Set up Selenium options
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument(f"user-agent={get_random_user_agent()}")
    
    service = Service(chrome_driver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)
    
    driver.get(url)
    html = driver.page_source
    driver.quit()
    
    soup = BeautifulSoup(html, 'html.parser')
    return soup

with open('Vietnam-housing-data/data collection scripts/product_links.txt', 'r') as file:
    links = [line.strip() for line in file.readlines()]

# Set up Excel file and sheet
file_path = 'Vietnam-housing-data/vietnam_housing_dataset.xlsx'
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.append([
        'Address', 'Area (m^2)', 'Frontage (m)', 'Access Road (m)', 
        'House direction', 'Balcony direction', 'No. of floors', 
        'No. of bedrooms', 'No. of bathrooms', 'Legal status', 
        'Furniture state', 'Price'
    ])
    wb.save(file_path)

wb = load_workbook(file_path)
ws = wb.active

# Loop through each link and write row by row to Excel
for product_count, url in enumerate(links, 1):
    print(f"Checking specs for product {product_count}")
    try:
        soup = get_url(url)
        specs = soup.find('div', class_='re__pr-specs-content js__other-info')
        if specs:
            address = soup.find('span', class_='re__pr-short-description js__pr-address').text.strip()
            features = specs.find_all('div', class_='re__pr-specs-content-item')
            feature_dict = {feature.find('span', class_='re__pr-specs-content-item-title').text.strip(): 
                            feature.find('span', class_='re__pr-specs-content-item-value').text.strip() for feature in features}
            row_data = [
                address,
                feature_dict.get('Diện tích'),
                feature_dict.get('Mặt tiền'),
                feature_dict.get('Đường vào'),
                feature_dict.get('Hướng nhà'),
                feature_dict.get('Hướng ban công'),
                feature_dict.get('Số tầng'),
                feature_dict.get('Số phòng ngủ'),
                feature_dict.get('Số toilet'),
                feature_dict.get('Pháp lý'),
                feature_dict.get('Nội thất'),
                feature_dict.get('Mức giá')
            ]
            ws.append(row_data)
            wb.save(file_path)
    except Exception as e:
        print(f"An error occurred during crawling: {e}")
    print("-" * 40)

print(f"Data has been successfully written to '{file_path}'")
